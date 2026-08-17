"""
Excel / PDF export helpers for admin data.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from fpdf import FPDF


def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def rows_to_xlsx(
    title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Export")[:31]
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(["" if v is None else v for v in row])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def rows_to_pdf(
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    landscape: bool = True,
) -> bytes:
    pdf = FPDF(orientation="L" if landscape else "P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(
        0,
        6,
        f"Generated (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(2)

    col_count = max(len(headers), 1)
    page_w = pdf.w - pdf.l_margin - pdf.r_margin
    col_w = page_w / col_count

    def _cell(text: Any, header: bool = False) -> None:
        pdf.set_font("Helvetica", "B" if header else "", 7)
        s = str("" if text is None else text)
        if len(s) > 40:
            s = s[:37] + "..."
        pdf.cell(col_w, 6, s, border=1)

    for h in headers:
        _cell(h, header=True)
    pdf.ln()
    for row in rows:
        for i in range(col_count):
            val = row[i] if i < len(row) else ""
            _cell(val)
        pdf.ln()

    out = pdf.output()
    if isinstance(out, (bytes, bytearray)):
        return bytes(out)
    return str(out).encode("latin-1", errors="replace")


def filename(prefix: str, ext: str) -> str:
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in prefix)
    return f"{safe}_{_ts()}.{ext.lstrip('.')}"
